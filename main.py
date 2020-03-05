#!/usr/bin/python
import os
import requests
from lxml import html
from openpyxl import Workbook, styles
from constants import headers

import sys


def main(args):

    # Setting the order filename
    order_filename = args.pop() if len(args) > 0 else 'sample'

    # Opening the import file with the order
    # Format: First line is the name of the buyer and then the item list (quantity url_item)
    import_file = 'imports/{}.txt'.format(order_filename)
    order_txt = open(import_file, 'r')

    # Creating the Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "PEDIDO_{}".format(order_filename)
    # Setting style set to use
    header_font = styles.Font(name='Arial', size=9, color="225a9d", bold=True)
    align_center = styles.Alignment(horizontal='center', vertical='center')

    # Creating worksheet headers
    row = 1
    col = 1
    for width, text in headers:
        ws.cell(row=row, column=col).alignment = align_center
        ws.cell(row=row, column=col).value = str(text or '')
        ws.cell(row=row, column=col).font = header_font
        ws.column_dimensions[chr(64+col)].width = width
        col += 1

    # Reading import file
    buyer_name = order_txt.readline()
    for item in order_txt:
        try:
            row += 1
            col = 1
            quantity, url = item.split(' ')
            quantity = int(quantity)
            # Making request to the item url
            page = requests.get(url)
            tree = html.fromstring(page.content)
            # Web scrapping info
            item = tree.xpath('//div[@id="prodContainer"]/div[2]/div[1]/h1[1]/text()')[0]
            price_text = tree.xpath('//div[contains(@class, "buyBox")][1]/div[1]/div[1]/div[1]/span/text()')[1]
            currency = price_text[0]
            price = float(price_text[1:])
            item_parts = item.split(' - ')
            item_parts_len = len(item_parts)
            item_name = item_parts[0] if item_parts_len == 3 else ' - '.join(item_parts[0:item_parts_len-2])
            set_number = item_parts[-2]
            rarity = item_parts[-1]
            expansion = tree.xpath('//div[@id="prodContainer"]/div[2]/div[2]/div[1]/a[1]/h2[1]/text()')[0]
            condition = 'NM'
            subtotal = quantity * price
            # Writing the data on the worksheet
            data = (buyer_name, quantity, item, expansion, set_number, rarity, condition, price, subtotal, url)
            for i, value in enumerate(data):
                ws.cell(row=row, column=i+1).value = value
        except ValueError as e:
            print(e)
            pass
    # Closing import file
    order_txt.close()

    # Creating the export file
    export_file = 'exports/{}.xlsx'.format(order_filename)
    # Overwriting if already exists
    if os.path.exists(export_file):
        os.remove(export_file)
    wb.save(filename=export_file)

    print('Done!')
    return True


if __name__ == "__main__":
    main(sys.argv[1:])
